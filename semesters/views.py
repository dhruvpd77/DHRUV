from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.core.files.base import ContentFile
from .models import Semester, Subject, Question, Unit
import pandas as pd
import openpyxl
from io import BytesIO
import os
import csv
from PIL import Image

@staff_member_required
def admin_dashboard(request):
    semesters = Semester.objects.all()
    subjects = Subject.objects.all()
    questions_count = Question.objects.count()
    
    context = {
        'semesters': semesters,
        'subjects': subjects,
        'questions_count': questions_count,
    }
    return render(request, 'semesters/admin_dashboard.html', context)

@staff_member_required
def create_semester(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        
        if name:
            Semester.objects.create(name=name, description=description)
            messages.success(request, f'Semester "{name}" created successfully!')
            return redirect('semesters:admin_dashboard')
        else:
            messages.error(request, 'Semester name is required.')
    
    return render(request, 'semesters/create_semester.html')

@staff_member_required
def edit_semester(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    
    if request.method == 'POST':
        semester.name = request.POST.get('name', semester.name)
        semester.description = request.POST.get('description', semester.description)
        semester.save()
        messages.success(request, f'Semester "{semester.name}" updated successfully!')
        return redirect('semesters:manage_semesters')
    
    return render(request, 'semesters/edit_semester.html', {'semester': semester})

@staff_member_required
def delete_semester(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    
    if request.method == 'POST':
        semester_name = semester.name
        semester.delete()
        messages.success(request, f'Semester "{semester_name}" deleted successfully!')
        return redirect('semesters:manage_semesters')
    
    # Get counts
    subjects_count = semester.subjects.count()
    questions_count = sum(subject.questions.count() for subject in semester.subjects.all())
    
    return render(request, 'semesters/delete_semester.html', {
        'semester': semester,
        'subjects_count': subjects_count,
        'questions_count': questions_count,
    })

@staff_member_required
def edit_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    semesters = Semester.objects.all()
    
    if request.method == 'POST':
        subject.name = request.POST.get('name', subject.name)
        subject.code = request.POST.get('code', subject.code)
        semester_id = request.POST.get('semester')
        if semester_id:
            subject.semester = get_object_or_404(Semester, id=semester_id)
        subject.save()
        messages.success(request, f'Subject "{subject.name}" updated successfully!')
        return redirect('semesters:manage_subjects', semester_id=subject.semester.id)
    
    return render(request, 'semesters/edit_subject.html', {
        'subject': subject,
        'semesters': semesters
    })

@staff_member_required
def delete_subject(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    semester_id = subject.semester.id
    
    if request.method == 'POST':
        subject_name = subject.name
        subject.delete()
        messages.success(request, f'Subject "{subject_name}" deleted successfully!')
        return redirect('semesters:manage_subjects', semester_id=semester_id)
    
    questions_count = subject.questions.count()
    
    return render(request, 'semesters/delete_subject.html', {
        'subject': subject,
        'questions_count': questions_count,
    })

@staff_member_required
def create_subject(request):
    semesters = Semester.objects.all()
    
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code', '')
        semester_id = request.POST.get('semester')
        
        if name and semester_id:
            semester = get_object_or_404(Semester, id=semester_id)
            Subject.objects.create(name=name, code=code, semester=semester)
            messages.success(request, f'Subject "{name}" created successfully!')
            return redirect('semesters:admin_dashboard')
        else:
            messages.error(request, 'Subject name and semester are required.')
    
    return render(request, 'semesters/create_subject.html', {'semesters': semesters})

@staff_member_required
def upload_questions(request):
    subjects = Subject.objects.all()
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        subject_id = request.POST.get('subject')
        
        if not excel_file:
            messages.error(request, 'Please select an Excel file.')
            return render(request, 'semesters/upload_questions.html', {'subjects': subjects})
        
        if not subject_id:
            messages.error(request, 'Please select a subject.')
            return render(request, 'semesters/upload_questions.html', {'subjects': subjects})
        
        subject = get_object_or_404(Subject, id=subject_id)
        
        try:
            # Save uploaded file temporarily
            file_path = f'/tmp/{excel_file.name}' if os.name != 'nt' else f'C:\\temp\\{excel_file.name}'
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            with open(file_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)
            
            # Try to extract images, but continue even if it fails
            images_dict = {}
            try:
                # Load workbook with openpyxl to extract images
                # Use data_only=True to avoid formula issues, read_only=False for images
                wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=False)
                ws = wb.active
                
                # Extract all images from the worksheet and map them to cells
                if hasattr(ws, '_images'):
                    for image in ws._images:
                        # Get the cell anchor (where image is placed)
                        try:
                            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                                row_idx = image.anchor._from.row + 1  # openpyxl is 0-indexed
                                col_idx = image.anchor._from.col + 1
                                
                                # Save image to BytesIO
                                img_data = BytesIO(image._data())
                                images_dict[row_idx] = img_data
                        except Exception as img_error:
                            # Skip this image if there's an error
                            continue
            except Exception as openpyxl_error:
                # If openpyxl fails, continue without images
                pass  # Will show message later
            
            # Try multiple engines to read the Excel file
            df = None
            engines_to_try = ['openpyxl', 'xlrd', None]  # None lets pandas choose
            last_error = None
            
            for engine in engines_to_try:
                try:
                    if engine:
                        df = pd.read_excel(file_path, dtype=str, keep_default_na=False, engine=engine)
                    else:
                        df = pd.read_excel(file_path, dtype=str, keep_default_na=False)
                    break  # Success! Exit loop
                except Exception as e:
                    last_error = e
                    continue  # Try next engine
            
            # If pandas failed, try reading cell-by-cell with openpyxl directly (LAST RESORT)
            if df is None:
                # Try multiple openpyxl strategies
                openpyxl_strategies = [
                    # Strategy 1: read_only mode (most forgiving)
                    {'read_only': True, 'data_only': True, 'keep_vba': False},
                    # Strategy 2: read_only without data_only
                    {'read_only': True, 'data_only': False, 'keep_vba': False},
                    # Strategy 3: Normal mode with repair
                    {'read_only': False, 'data_only': True, 'keep_vba': False},
                    # Strategy 4: Normal mode without data_only
                    {'read_only': False, 'data_only': False, 'keep_vba': False},
                ]
                
                for strategy_idx, strategy in enumerate(openpyxl_strategies):
                    try:
                        if strategy_idx == 0:
                            messages.warning(request, 'File has severe XML corruption. Attempting direct cell-by-cell reading...')
                        
                        # Try to open with current strategy
                        wb_readonly = openpyxl.load_workbook(file_path, **strategy)
                        ws_readonly = wb_readonly.active
                        
                        # Read data manually row by row
                        data_rows = []
                        headers = None
                        max_rows = 10000  # Safety limit
                        
                        for idx, row in enumerate(ws_readonly.iter_rows(values_only=True)):
                            if idx >= max_rows:
                                break
                            
                            if idx == 0:
                                # First row is headers
                                headers = [str(cell) if cell is not None else '' for cell in row]
                            else:
                                # Data rows
                                row_data = [str(cell) if cell is not None else '' for cell in row]
                                if any(row_data):  # Skip completely empty rows
                                    data_rows.append(row_data)
                        
                        # Create DataFrame from manually read data
                        if headers and data_rows:
                            df = pd.DataFrame(data_rows, columns=headers)
                            messages.success(request, f'Successfully read {len(df)} rows using cell-by-cell method (strategy {strategy_idx + 1})!')
                            wb_readonly.close()
                            break  # Success! Exit loop
                        
                        wb_readonly.close()
                    
                    except Exception as cell_error:
                        # Try next strategy
                        continue
                
                # If all openpyxl strategies failed, try CSV conversion as absolute last resort
                if df is None:
                    try:
                        messages.warning(request, 'All Excel reading methods failed. Attempting CSV conversion workaround...')
                        
                        # Try to read as CSV (if file can be opened as text)
                        csv_data = []
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            # Try to detect if it's actually a CSV or TSV
                            sample = f.read(1024)
                            f.seek(0)
                            
                            # Try comma-separated first
                            try:
                                reader = csv.reader(f)
                                for row in reader:
                                    if row:  # Skip empty rows
                                        csv_data.append(row)
                                if len(csv_data) > 1:  # At least header + 1 row
                                    headers = [str(h) for h in csv_data[0]]
                                    data_rows = [row for row in csv_data[1:] if any(row)]
                                    if headers and data_rows:
                                        df = pd.DataFrame(data_rows, columns=headers)
                                        messages.success(request, f'Successfully read {len(df)} rows using CSV fallback method!')
                            except:
                                pass
                    except Exception as csv_error:
                        pass
            
            # If still no data, provide helpful error message with actionable steps
            if df is None:
                # Create a more user-friendly error message
                error_msg = (
                    "❌ Cannot read Excel file due to severe XML corruption.\n\n"
                    "🔧 QUICK FIX (Recommended - Takes 2 minutes):\n"
                    "1. Open the corrupted file in Microsoft Excel\n"
                    "2. Press Ctrl+A to select all data\n"
                    "3. Press Ctrl+C to copy\n"
                    "4. Close the corrupted file\n"
                    "5. Create a NEW blank Excel workbook\n"
                    "6. Press Ctrl+V to paste\n"
                    "7. Save as new .xlsx file (e.g., 'Fixed_File.xlsx')\n"
                    "8. Upload the new file\n\n"
                    "💡 Alternative: Save as CSV first, then convert back to Excel\n"
                    "1. Open in Excel → File → Save As → CSV (Comma delimited)\n"
                    "2. Close and reopen the CSV file\n"
                    "3. Save As → Excel Workbook (.xlsx)\n"
                    "4. Upload the new file\n\n"
                    "🛠️ Or run fix script: python fix_corrupted_excel.py \"your_file.xlsx\"\n\n"
                    "⚠️ The file structure is too corrupted for automatic repair."
                )
                raise Exception(error_msg)
            
            # Expected columns:
            # unit_number | question_text | MCQ Answer | option A | option B | option C | option D | Added By | Verified By
            
            questions_created = 0
            skipped_count = 0
            images_extracted = 0
            
            for index, row in df.iterrows():
                try:
                    # Excel row number (pandas index + 2 for header + 0-indexing)
                    excel_row = index + 2
                    
                    # Extract unit number from "unit_number" column
                    unit = row.get('unit_number', '1')
                    if not unit or unit == '' or unit == 'nan':
                        unit = 1
                    else:
                        try:
                            unit = int(float(unit))
                        except:
                            unit = 1
                    
                    # Get values and preserve whitespace/indentation
                    question_text = row.get('question_text', '')
                    mcq_answer_raw = str(row.get('MCQ Answer', '')).strip()  # Keep original for text matching
                    mcq_answer = mcq_answer_raw.upper()  # Uppercase for letter matching
                    option_a = row.get('option A', '')
                    option_b = row.get('option B', '')
                    option_c = row.get('option C', '')
                    option_d = row.get('option D', '')
                    added_by = row.get('Added By', '')
                    verified_by = row.get('Verified By', '')
                    
                    # Clean 'nan' strings but preserve actual content
                    def clean_value(val):
                        if val == 'nan' or val == '':
                            return ''
                        return val
                    
                    # Clean all values
                    question_text = clean_value(question_text)
                    option_a = clean_value(option_a)
                    option_b = clean_value(option_b)
                    option_c = clean_value(option_c)
                    option_d = clean_value(option_d)
                    
                    # ONLY EXTRACT MCQs - Skip if question or ANY option is empty
                    if not question_text or question_text.strip() == '':
                        skipped_count += 1
                        continue
                    
                    if not option_a or option_a.strip() == '':
                        skipped_count += 1
                        continue
                    
                    if not option_b or option_b.strip() == '':
                        skipped_count += 1
                        continue
                    
                    if not option_c or option_c.strip() == '':
                        skipped_count += 1
                        continue
                    
                    if not option_d or option_d.strip() == '':
                        skipped_count += 1
                        continue
                    
                    # Convert MCQ Answer to letter format (A, B, C, D)
                    # Support both formats:
                    # 1. Existing format: "A", "B", "C", "D", "Option A", "A)", etc.
                    # 2. New format: Actual text from options (e.g., "Uniform Resource Locator")
                    correct_answer = None
                    
                    # Helper function to normalize text for comparison (case-insensitive, normalize whitespace)
                    def normalize_text(text):
                        if not text:
                            return ''
                        # Strip and normalize multiple spaces to single space
                        normalized = ' '.join(str(text).strip().split())
                        return normalized
                    
                    # First, try existing format (letter-based matching)
                    if mcq_answer in ['A', 'B', 'C', 'D']:
                        correct_answer = mcq_answer
                    elif mcq_answer.startswith('A'):
                        correct_answer = 'A'
                    elif mcq_answer.startswith('B'):
                        correct_answer = 'B'
                    elif mcq_answer.startswith('C'):
                        correct_answer = 'C'
                    elif mcq_answer.startswith('D'):
                        correct_answer = 'D'
                    
                    # If not found, try new format (text-based matching)
                    # Compare the MCQ answer text with each option text (case-insensitive, normalized whitespace)
                    if correct_answer is None:
                        mcq_answer_normalized = normalize_text(mcq_answer_raw).lower()
                        option_a_normalized = normalize_text(option_a).lower()
                        option_b_normalized = normalize_text(option_b).lower()
                        option_c_normalized = normalize_text(option_c).lower()
                        option_d_normalized = normalize_text(option_d).lower()
                        
                        # Case-insensitive exact match
                        if mcq_answer_normalized == option_a_normalized:
                            correct_answer = 'A'
                        elif mcq_answer_normalized == option_b_normalized:
                            correct_answer = 'B'
                        elif mcq_answer_normalized == option_c_normalized:
                            correct_answer = 'C'
                        elif mcq_answer_normalized == option_d_normalized:
                            correct_answer = 'D'
                    
                    # Default fallback if no match found
                    if correct_answer is None:
                        correct_answer = 'A'  # Default
                    
                    # Create question
                    question = Question.objects.create(
                        subject=subject,
                        unit=unit,
                        question_text=question_text,
                        option_a=option_a,
                        option_b=option_b,
                        option_c=option_c,
                        option_d=option_d,
                        correct_answer=correct_answer,
                        added_by=clean_value(added_by),
                        verified_by=clean_value(verified_by)
                    )
                    
                    # Check if this row has an image
                    if excel_row in images_dict:
                        try:
                            img_data = images_dict[excel_row]
                            img_data.seek(0)
                            
                            # Generate unique filename
                            filename = f'question_{question.id}_{excel_row}.png'
                            
                            # Save image to question
                            question.question_image.save(filename, ContentFile(img_data.read()), save=True)
                            images_extracted += 1
                        except Exception as img_error:
                            pass  # Continue even if image extraction fails
                    
                    questions_created += 1
                except Exception as e:
                    skipped_count += 1
                    continue
            
            # Clean up temp file
            try:
                os.remove(file_path)
            except:
                pass
            
            if questions_created > 0:
                msg = f'✅ {questions_created} MCQ questions uploaded successfully!'
                if images_extracted > 0:
                    msg += f' ({images_extracted} questions with images)'
                elif len(images_dict) == 0:
                    msg += ' ⚠️ No images were extracted (file may have XML issues or no images present)'
                if skipped_count > 0:
                    msg += f' (Skipped {skipped_count} incomplete rows)'
                messages.success(request, msg)
            else:
                messages.warning(request, f'No valid MCQ questions found. Skipped {skipped_count} rows. Make sure all 4 options (A, B, C, D) have values.')
            return redirect('semesters:admin_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            # Clean up temp file on error
            try:
                if 'file_path' in locals():
                    os.remove(file_path)
            except:
                pass
    
    return render(request, 'semesters/upload_questions.html', {'subjects': subjects})

@staff_member_required
def manage_semesters(request):
    semesters = Semester.objects.all()
    return render(request, 'semesters/manage_semesters.html', {'semesters': semesters})

@staff_member_required
def manage_subjects(request, semester_id):
    semester = get_object_or_404(Semester, id=semester_id)
    subjects = Subject.objects.filter(semester=semester)
    return render(request, 'semesters/manage_subjects.html', {
        'semester': semester,
        'subjects': subjects
    })

@staff_member_required
def manage_questions(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    unit_filter = request.GET.get('unit', '')
    
    if unit_filter:
        questions = Question.objects.filter(subject=subject, unit=unit_filter).order_by('id')
    else:
        questions = Question.objects.filter(subject=subject).order_by('unit', 'id')
    
    # Get all available units for this subject
    available_units = Question.objects.filter(subject=subject).values_list('unit', flat=True).distinct().order_by('unit')
    
    context = {
        'subject': subject,
        'questions': questions,
        'available_units': available_units,
        'selected_unit': unit_filter,
    }
    return render(request, 'semesters/manage_questions.html', context)

@staff_member_required
def edit_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    
    if request.method == 'POST':
        question.unit = request.POST.get('unit', question.unit)
        question.question_text = request.POST.get('question_text', question.question_text)
        question.option_a = request.POST.get('option_a', question.option_a)
        question.option_b = request.POST.get('option_b', question.option_b)
        question.option_c = request.POST.get('option_c', question.option_c)
        question.option_d = request.POST.get('option_d', question.option_d)
        question.correct_answer = request.POST.get('correct_answer', question.correct_answer)
        question.added_by = request.POST.get('added_by', question.added_by)
        question.verified_by = request.POST.get('verified_by', question.verified_by)
        
        # Handle image upload
        if 'question_image' in request.FILES:
            question.question_image = request.FILES['question_image']
        
        # Handle image removal
        if request.POST.get('remove_image') == 'on' and question.question_image:
            question.question_image.delete()
            question.question_image = None
        
        question.save()
        
        messages.success(request, 'Question updated successfully!')
        return redirect('semesters:manage_questions', subject_id=question.subject.id)
    
    return render(request, 'semesters/edit_question.html', {'question': question})

@staff_member_required
def delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    subject_id = question.subject.id
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted successfully!')
        return redirect('semesters:manage_questions', subject_id=subject_id)
    
    return render(request, 'semesters/delete_question.html', {'question': question})

@staff_member_required
def add_question(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == 'POST':
        unit = request.POST.get('unit', 1)
        question_text = request.POST.get('question_text', '')
        option_a = request.POST.get('option_a', '')
        option_b = request.POST.get('option_b', '')
        option_c = request.POST.get('option_c', '')
        option_d = request.POST.get('option_d', '')
        correct_answer = request.POST.get('correct_answer', 'A')
        added_by = request.POST.get('added_by', '')
        verified_by = request.POST.get('verified_by', '')
        
        # Validate all fields
        if question_text and option_a and option_b and option_c and option_d:
            question = Question.objects.create(
                subject=subject,
                unit=unit,
                question_text=question_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                correct_answer=correct_answer,
                added_by=added_by,
                verified_by=verified_by
            )
            
            # Handle image upload
            if 'question_image' in request.FILES:
                question.question_image = request.FILES['question_image']
                question.save()
            
            messages.success(request, 'Question added successfully!')
            return redirect('semesters:manage_questions', subject_id=subject.id)
        else:
            messages.error(request, 'All question fields and options are required!')
    
    return render(request, 'semesters/add_question.html', {'subject': subject})

@staff_member_required
def delete_unit_questions(request, subject_id, unit):
    subject = get_object_or_404(Subject, id=subject_id)
    questions = Question.objects.filter(subject=subject, unit=unit)
    question_count = questions.count()
    
    if request.method == 'POST':
        questions.delete()
        messages.success(request, f'Successfully deleted {question_count} questions from Unit {unit}!')
        return redirect('semesters:manage_questions', subject_id=subject.id)
    
    context = {
        'subject': subject,
        'unit': unit,
        'question_count': question_count,
        'questions': questions[:5]  # Show first 5 as preview
    }
    return render(request, 'semesters/delete_unit_questions.html', context)

@staff_member_required
def delete_all_questions(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    questions = Question.objects.filter(subject=subject)
    question_count = questions.count()
    
    # Get unit distribution for preview
    unit_distribution = {}
    for unit in questions.values_list('unit', flat=True).distinct().order_by('unit'):
        unit_distribution[unit] = questions.filter(unit=unit).count()
    
    if request.method == 'POST':
        questions.delete()
        messages.success(request, f'Successfully deleted all {question_count} questions from {subject.name}!')
        return redirect('semesters:manage_questions', subject_id=subject.id)
    
    context = {
        'subject': subject,
        'question_count': question_count,
        'unit_distribution': unit_distribution,
    }
    return render(request, 'semesters/delete_all_questions.html', context)


@staff_member_required
def manage_units(request, subject_id):
    """Manage units for a subject"""
    subject = get_object_or_404(Subject, id=subject_id)
    units = Unit.objects.filter(subject=subject).order_by('unit_number')
    
    # Get distinct units from questions
    question_units = Question.objects.filter(subject=subject).values_list('unit', flat=True).distinct().order_by('unit')
    
    # Get existing unit numbers
    existing_unit_numbers = set(units.values_list('unit_number', flat=True))
    
    return render(request, 'semesters/manage_units.html', {
        'subject': subject,
        'units': units,
        'question_units': question_units,
        'existing_unit_numbers': existing_unit_numbers,
    })


@staff_member_required
def create_unit(request, subject_id, unit_number):
    """Create unit information"""
    subject = get_object_or_404(Subject, id=subject_id)
    unit, created = Unit.objects.get_or_create(
        subject=subject,
        unit_number=unit_number,
        defaults={'title': '', 'topics': ''}
    )
    if created:
        messages.info(request, f'Created new Unit {unit_number}. Please add title and topics.')
    
    if request.method == 'POST':
        unit.title = request.POST.get('title', '')
        unit.topics = request.POST.get('topics', '')
        unit.save()
        messages.success(request, f'Unit {unit.unit_number} saved successfully!')
        return redirect('semesters:manage_units', subject_id=unit.subject.id)
    
    return render(request, 'semesters/edit_unit.html', {'unit': unit})


@staff_member_required
def edit_unit(request, unit_id):
    """Edit unit information"""
    unit = get_object_or_404(Unit, id=unit_id)
    
    if request.method == 'POST':
        unit.title = request.POST.get('title', '')
        unit.topics = request.POST.get('topics', '')
        unit.save()
        messages.success(request, f'Unit {unit.unit_number} saved successfully!')
        return redirect('semesters:manage_units', subject_id=unit.subject.id)
    
    return render(request, 'semesters/edit_unit.html', {'unit': unit})
